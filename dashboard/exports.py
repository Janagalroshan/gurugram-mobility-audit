"""
Builds, on a button click:
  - a 10-sheet Excel workbook (Cover, Ranking, Hourly Medians, Direction
    Asymmetry, Reliability, Coverage, FAIL Log, Distance Drift, Methodology,
    Raw Observations) ready to staple into the audit report.
  - a zip of chart images at ~1600x1000 px for the Word report.
"""

from __future__ import annotations

import io
import zipfile
from datetime import datetime

import pandas as pd
import pytz
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

import metrics
import ui

# All "as of now" timestamps written into audit outputs must read in IST,
# never in the export server's own local/UTC clock (Streamlit Cloud runs
# on UTC), so every report matches the IST timestamps already in the data.
IST = pytz.timezone("Asia/Kolkata")


def _write_df(ws, df: pd.DataFrame, title: str | None = None):
    row0 = 1
    if title:
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
        row0 = 3
    if df is None or df.empty:
        ws.cell(row=row0, column=1, value="No data available for this audit window.")
        return
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=row0):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == row0:
                cell.font = Font(bold=True)


def build_excel_workbook(df: pd.DataFrame, coverage: dict, raw_all: pd.DataFrame,
                          fingerprint: str) -> bytes:
    wb = Workbook()

    # 1. Cover
    ws = wb.active
    ws.title = "Cover"
    ws.cell(row=1, column=1, value=f"{ui.AUDIT_TITLE}").font = Font(bold=True, size=16)
    ws.cell(row=2, column=1, value=ui.AUDIT_OFFICE)
    ws.cell(row=4, column=1, value=f"Audit window: see Methodology sheet")
    ws.cell(row=5, column=1, value=f"Report generated: {datetime.now(IST).strftime('%Y-%m-%d %H:%M')} IST")
    ws.cell(row=6, column=1, value=f"Data fingerprint (MD5): {fingerprint}")

    # 2. Ranking
    phci_df = metrics.phci(df, weekday_only=True)
    _write_df(wb.create_sheet("Ranking"), phci_df, "Peak-Hour Congestion Index — Ranking")

    # 3. Hourly Medians
    hourly = metrics.hourly_median_cr(df)
    _write_df(wb.create_sheet("Hourly Medians"), hourly, "Hourly Median Congestion Ratio")

    # 4. Direction Asymmetry
    asym = metrics.direction_asymmetry(df)
    _write_df(wb.create_sheet("Direction Asymmetry"), asym, "AM vs PM Direction Asymmetry")

    # 5. Reliability
    rel = metrics.reliability(df)
    _write_df(wb.create_sheet("Reliability"), rel, "Buffer Time Index (BTI) and Coefficient of Variation (CV)")

    # 6. Coverage
    cov_df = pd.DataFrame([coverage])
    _write_df(wb.create_sheet("Coverage"), cov_df, "Data Coverage Summary")

    # 7. FAIL Log
    if raw_all is not None and not raw_all.empty:
        fails = raw_all[raw_all["api_status"] != "OK"]
    else:
        fails = pd.DataFrame()
    _write_df(wb.create_sheet("FAIL Log"), fails, "Failed API Readings (excluded from analysis, shown for transparency)")

    # 8. Distance Drift (rows where recorded distance varies meaningfully run-to-run)
    if not df.empty and "est_distance_km" in df.columns:
        drift = df.groupby(["corridor_id", "corridor_name"]).agg(
            est_distance_km=("est_distance_km", "first"),
            observed_distance_m_mean=("distance_m", "mean"),
            observed_distance_m_std=("distance_m", "std"),
        ).reset_index()
    else:
        drift = pd.DataFrame()
    _write_df(wb.create_sheet("Distance Drift"), drift, "Recorded vs Observed Distance (route-drift check)")

    # 9. Methodology
    meth_ws = wb.create_sheet("Methodology")
    meth_ws.cell(row=1, column=1, value="Methodology").font = Font(bold=True, size=13)
    lines = [
        "Congestion Ratio (CR) = duration_traffic_s / duration_freeflow_s",
        "Hourly value = median of CR over the hour (median, not mean)",
        "PHCI = max over peak hours, weekdays, of the hourly median CR (worse direction)",
        "ADCI = mean of hourly median CR across 06:00-22:00, weekdays",
        "BTI = (p95 peak travel time - median peak travel time) / median",
        "CV = std(peak travel time) / mean(peak travel time)",
        f"Peak hours used: AM {metrics.PEAK_HOURS_AM}, PM {metrics.PEAK_HOURS_PM} "
        f"(anchor: local government office hours — confirm/replace with the Haryana "
        f"Government's notified circular before finalising the report).",
        f"National MoUD SLB toggle window: AM {metrics.NATIONAL_SLB_AM}, PM {metrics.NATIONAL_SLB_PM}",
        "Gating thresholds (min. / stable observations) — see Appendix D of the build guide:",
    ]
    for key, g in metrics.GATES.items():
        lines.append(f"  {key}: preliminary >= {g['preliminary']}, stable >= {g['stable']}")
    lines.append(f"Data fingerprint (MD5): {fingerprint}")
    for i, line in enumerate(lines, start=3):
        meth_ws.cell(row=i, column=1, value=line)

    # 10. Raw Observations
    _write_df(wb.create_sheet("Raw Observations"), df, "Cleaned Raw Observations (audit window)")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_chart_zip(figures: dict) -> bytes:
    """figures: {filename_without_ext: plotly Figure}. Requires kaleido."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, fig in figures.items():
            try:
                png_bytes = fig.to_image(format="png", width=1600, height=1000, scale=1)
                zf.writestr(f"{name}.png", png_bytes)
            except Exception as e:  # noqa: BLE001
                zf.writestr(f"{name}_ERROR.txt", f"Could not render: {e}")
    return buf.getvalue()
